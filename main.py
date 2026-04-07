#!/usr/bin/env python3
"""
ProxiAlpha - Main Entry Point
=========================================
Modular pullback trading system with multiple strategies,
backtesting, paper trading, live trading, and Claude AI integration.

Usage:
    python main.py --mode scan          # Scan watchlist for signals
    python main.py --mode backtest      # Run backtest on historical data
    python main.py --mode paper         # Start paper trading
    python main.py --mode live          # Start live trading (Alpaca)
    python main.py --mode ai-signals    # Get AI-powered signals from Claude
    python main.py --mode excel         # Generate Excel tracker
    python main.py --mode all           # Run full analysis + generate all outputs
"""
import argparse
import json
import sys
import os
import yaml
from pathlib import Path

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from core.data_engine import fetch_all_watchlist, calculate_technical_indicators, get_full_analysis
from core.config import WATCHLIST, DEFAULT_CAPITAL
from strategies.strategy_manager import StrategyManager
from strategies.dip_buyer import DipBuyerStrategy
from strategies.technical import TechnicalStrategy
from strategies.dca import DCAStrategy
from strategies.custom_rules import CustomRulesStrategy
from strategies.ai_strategy import AIStrategy
from backtesting.engine import BacktestEngine
from paper_trading.simulator import PaperTrader


def load_yaml_config(filename):
    """Load a YAML config file."""
    path = Path(__file__).parent / filename
    if path.exists():
        with open(path) as f:
            return yaml.safe_load(f)
    return {}


def build_strategy_manager(config=None):
    """Build and configure the strategy manager from config."""
    manager = StrategyManager()

    if config and 'strategies' in config:
        strats = config['strategies']

        if strats.get('dip_buyer', {}).get('enabled', True):
            params = strats['dip_buyer'].get('parameters', {})
            manager.register_strategy(DipBuyerStrategy(
                weight=strats['dip_buyer'].get('weight', 1.0),
                params=params if params else None,
            ))

        if strats.get('technical', {}).get('enabled', True):
            manager.register_strategy(TechnicalStrategy(
                weight=strats['technical'].get('weight', 1.0),
            ))

        if strats.get('dca', {}).get('enabled', True):
            params = strats['dca'].get('parameters', {})
            manager.register_strategy(DCAStrategy(
                weight=strats['dca'].get('weight', 1.0),
                params=params if params else None,
            ))

        if strats.get('custom_rules', {}).get('enabled', True):
            manager.register_strategy(CustomRulesStrategy(
                weight=strats['custom_rules'].get('weight', 1.0),
            ))

        if strats.get('ai_claude', {}).get('enabled', False):
            ai_params = strats['ai_claude'].get('parameters', {})
            manager.register_strategy(AIStrategy(
                weight=strats['ai_claude'].get('weight', 1.5),
                params=ai_params,
            ))
    else:
        # Default: register all strategies
        manager.register_strategy(DipBuyerStrategy(weight=1.2))
        manager.register_strategy(TechnicalStrategy(weight=1.0))
        manager.register_strategy(DCAStrategy(weight=0.8))
        manager.register_strategy(CustomRulesStrategy(weight=0.9))

    return manager


def run_scan(manager, data):
    """Scan all stocks and print signals."""
    print("\n" + "=" * 70)
    print("  PULLBACK WATCHLIST SCAN")
    print("=" * 70)

    results = manager.scan_all_tickers(data)
    print(results.to_string(index=False))

    # Highlight actionable signals
    buys = results[results['Signal'].isin(['BUY', 'STRONG_BUY'])]
    if not buys.empty:
        print(f"\n  >>> {len(buys)} BUY SIGNALS <<<")
        for _, row in buys.iterrows():
            print(f"    {row['Ticker']:6s} | {row['Signal']:12s} | Confidence: {row['Confidence']:.0%} | Target: ${row['Target'] or 'N/A'}")

    return results


def run_backtest(manager, data, initial_capital=DEFAULT_CAPITAL):
    """Run backtest and print results."""
    print("\n" + "=" * 70)
    print("  BACKTESTING")
    print("=" * 70)

    engine = BacktestEngine(manager, initial_capital)
    results = engine.run(data)
    summary = engine.get_summary()

    for key, value in summary.items():
        print(f"  {key:25s}: {value}")

    # Save results
    output_dir = Path(__file__).parent / "data"
    output_dir.mkdir(exist_ok=True)

    if not results['equity_curve'].empty:
        results['equity_curve'].to_csv(output_dir / "backtest_equity.csv", index=False)
    if not results['trade_log'].empty:
        results['trade_log'].to_csv(output_dir / "backtest_trades.csv", index=False)

    with open(output_dir / "backtest_summary.json", 'w') as f:
        json.dump(summary, f, indent=2)

    print(f"\n  Results saved to {output_dir}/")
    return summary


def run_paper_trading(manager):
    """Start paper trading session."""
    print("\n" + "=" * 70)
    print("  PAPER TRADING")
    print("=" * 70)

    trader = PaperTrader(manager, state_file=str(Path(__file__).parent / "data" / "paper_trading_state.json"))

    # Run scan
    signals = trader.run_scan()
    print(f"\n  Scanned {len(signals)} stocks")

    # Show proposed trades
    proposed = trader.auto_execute(signals, require_confirmation=True)
    if proposed:
        print(f"\n  Proposed trades:")
        for trade in proposed:
            print(f"    {trade['action']:4s} {trade['ticker']:6s} | {trade.get('shares', 0)} shares @ ${trade.get('price', 0):.2f}")
    else:
        print("  No trades proposed at this time.")

    # Show performance
    perf = trader.get_performance()
    if perf:
        print(f"\n  Portfolio: ${perf.get('equity', 0):,.2f} ({perf.get('total_return_pct', 0):+.1f}%)")
        print(f"  Cash: ${perf.get('cash', 0):,.2f} | Positions: {perf.get('num_positions', 0)}")

    return proposed


def run_ai_signals(manager, data):
    """Generate AI signals (requires Anthropic API key)."""
    print("\n" + "=" * 70)
    print("  CLAUDE AI SIGNAL GENERATION")
    print("=" * 70)

    ai_config = load_yaml_config("config_ai_integration.yaml")
    api_key = ai_config.get('anthropic', {}).get('api_key') or os.environ.get('ANTHROPIC_API_KEY')

    if not api_key:
        print("  No API key found. Set in config_ai_integration.yaml or ANTHROPIC_API_KEY env var.")
        print("  Falling back to rule-based strategies only.\n")
        return run_scan(manager, data)

    # Enable AI strategy
    ai_strat = AIStrategy(weight=1.5, params={'api_key': api_key, 'use_api': True})
    manager.register_strategy(ai_strat)

    results = manager.scan_all_tickers(data)
    print(results.to_string(index=False))
    return results


def generate_excel(analysis_df, data):
    """Generate Excel tracker spreadsheet."""
    print("\n  Generating Excel tracker...")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # --- Sheet 1: Pullback Analysis ---
    ws = wb.active
    ws.title = "Pullback Analysis"

    headers = ['Ticker', 'Sector', 'ATH ($)', 'Low ($)', 'Current ($)',
               'Drawdown %', 'Recovery %', 'Upside to ATH %', 'RSI',
               'Above SMA20', 'Above SMA50', 'MACD Bullish', 'Vol Ratio']

    header_fill = PatternFill('solid', fgColor='1F2937')
    header_font = Font(bold=True, color='FFFFFF', size=11, name='Arial')
    blue_font = Font(color='0000FF', name='Arial', size=10)
    black_font = Font(name='Arial', size=10)
    green_font = Font(color='008000', name='Arial', size=10)
    red_font = Font(color='FF0000', name='Arial', size=10)
    thin_border = Border(
        bottom=Side(style='thin', color='E5E7EB')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    row = 2
    for _, stock in analysis_df.iterrows():
        ws.cell(row=row, column=1, value=stock['ticker']).font = Font(bold=True, color='3B82F6', name='Arial', size=10)
        ws.cell(row=row, column=2, value=stock.get('sector', '')).font = black_font
        ws.cell(row=row, column=3, value=stock.get('all_time_high', 0)).font = blue_font
        ws.cell(row=row, column=3).number_format = '$#,##0'
        ws.cell(row=row, column=4, value=stock.get('pullback_low', 0)).font = blue_font
        ws.cell(row=row, column=4).number_format = '$#,##0'
        ws.cell(row=row, column=5, value=stock.get('current_price', 0)).font = Font(bold=True, name='Arial', size=10)
        ws.cell(row=row, column=5).number_format = '$#,##0.00'

        dd = stock.get('drawdown_pct', 0)
        ws.cell(row=row, column=6, value=dd / 100 if dd else 0).font = red_font
        ws.cell(row=row, column=6).number_format = '0.0%'

        rec = stock.get('recovery_pct', 0)
        ws.cell(row=row, column=7, value=rec / 100 if rec else 0).font = green_font
        ws.cell(row=row, column=7).number_format = '0.0%'

        upside = stock.get('upside_to_ath_pct', 0)
        ws.cell(row=row, column=8, value=upside / 100 if upside else 0).font = green_font
        ws.cell(row=row, column=8).number_format = '0.0%'

        rsi = stock.get('rsi', 0)
        rsi_font = green_font if rsi and rsi < 30 else (red_font if rsi and rsi > 70 else black_font)
        ws.cell(row=row, column=9, value=rsi).font = rsi_font

        ws.cell(row=row, column=10, value='Yes' if stock.get('above_sma20') else 'No').font = black_font
        ws.cell(row=row, column=11, value='Yes' if stock.get('above_sma50') else 'No').font = black_font
        ws.cell(row=row, column=12, value='Yes' if stock.get('macd_bullish') else 'No').font = black_font
        ws.cell(row=row, column=13, value=stock.get('vol_ratio', 0)).font = black_font
        ws.cell(row=row, column=13).number_format = '0.00'

        for col in range(1, 14):
            ws.cell(row=row, column=col).border = thin_border
        row += 1

    # Column widths
    widths = [10, 18, 10, 10, 12, 12, 12, 15, 8, 12, 12, 14, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = 'A2'

    # --- Sheet 2: Summary Formulas ---
    ws2 = wb.create_sheet("Summary")
    ws2['A1'] = 'Pullback Trading Summary'
    ws2['A1'].font = Font(bold=True, size=14, name='Arial')

    num_stocks = len(analysis_df)
    ws2['A3'] = 'Total Stocks'
    ws2['B3'] = f'=COUNTA(\'Pullback Analysis\'!A2:A{num_stocks+1})'
    ws2['A4'] = 'Avg Drawdown'
    ws2['B4'] = f'=AVERAGE(\'Pullback Analysis\'!F2:F{num_stocks+1})'
    ws2['B4'].number_format = '0.0%'
    ws2['A5'] = 'Avg RSI'
    ws2['B5'] = f'=AVERAGE(\'Pullback Analysis\'!I2:I{num_stocks+1})'
    ws2['B5'].number_format = '0.0'
    ws2['A6'] = 'Stocks RSI < 30 (Oversold)'
    ws2['B6'] = f'=COUNTIF(\'Pullback Analysis\'!I2:I{num_stocks+1},"<30")'
    ws2['A7'] = 'Avg Upside to ATH'
    ws2['B7'] = f'=AVERAGE(\'Pullback Analysis\'!H2:H{num_stocks+1})'
    ws2['B7'].number_format = '0.0%'

    for r in range(3, 8):
        ws2[f'A{r}'].font = Font(bold=True, name='Arial', size=10)
        ws2[f'B{r}'].font = Font(name='Arial', size=10)

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 15

    output_path = Path(__file__).parent / "data" / "pullback_tracker.xlsx"
    output_path.parent.mkdir(exist_ok=True)
    wb.save(str(output_path))
    print(f"  Excel tracker saved to {output_path}")
    return str(output_path)


def main():
    parser = argparse.ArgumentParser(description='ProxiAlpha Trading Platform')
    parser.add_argument('--mode', choices=['scan', 'backtest', 'paper', 'live', 'ai-signals', 'excel', 'all'],
                        default='all', help='Operating mode')
    parser.add_argument('--capital', type=float, default=DEFAULT_CAPITAL, help='Initial capital')
    args = parser.parse_args()

    print("=" * 70)
    print("  PROXIALPHA TRADING PLATFORM")
    print("  AI-Powered Modular Trading System")
    print("=" * 70)

    # Load configs
    strat_config = load_yaml_config("config_strategies.yaml")
    trading_config = load_yaml_config("config_trading.yaml")

    # Build strategy manager
    manager = build_strategy_manager(strat_config)
    print(f"\n  Strategies loaded: {list(manager.strategies.keys())}")

    # Fetch data
    print("  Fetching market data...")
    analysis_df, enriched_data = get_full_analysis()
    print(f"  Loaded data for {len(enriched_data)} stocks\n")

    mode = args.mode

    if mode in ('scan', 'all'):
        run_scan(manager, enriched_data)

    if mode in ('backtest', 'all'):
        run_backtest(manager, enriched_data, args.capital)

    if mode in ('paper', 'all'):
        run_paper_trading(manager)

    if mode == 'live':
        from live_trading.alpaca_bot import AlpacaLiveTrader
        trader = AlpacaLiveTrader(manager)
        trader.run_loop(interval_minutes=60, auto_execute=False)

    if mode in ('ai-signals',):
        run_ai_signals(manager, enriched_data)

    if mode in ('excel', 'all'):
        generate_excel(analysis_df, enriched_data)

    if mode == 'all':
        # Export strategy config
        config_path = Path(__file__).parent / "data" / "current_strategy_config.json"
        with open(config_path, 'w') as f:
            f.write(manager.export_config())
        print(f"\n  Strategy config exported to {config_path}")

    print("\n" + "=" * 70)
    print("  DONE")
    print("=" * 70)


if __name__ == "__main__":
    main()
